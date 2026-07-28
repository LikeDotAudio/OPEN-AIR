"""Regenerate Yak/CommandList.csv (and .xlsx) from the YAK command tables.

The sheet is a REPORT, not a source. Every row is derived from a
`Yak/<Family>/<Model>/commands.json` entry — edit the table, run this, never the
other way round. The README said "edit commands.json and regenerate" without
shipping the regenerator; this is it.

One row per command. The `Returns` / `Return type` / `Return unit` /
`Return fields` group is the declared shape of a NAB's reply — nothing parses
replies yet (design audit §5.3), so those columns are the spec the receiver will
be built against rather than a description of anything running.

`Unverified` is the column the sheet did not have: a command swept out of a
manual and never sent to an instrument. That is 86% of the vocabulary, and a
sheet that stays quiet about it reads as 3600 working commands.

`Arguments` and `Instance params` split the SCPI placeholders two ways, and the
split is the whole reason a panel author reads this file. Arguments come from the
operator through sibling `Input/*` widgets; instance params (`<chan>`, `<slot>`)
are stamped per panel by build_instrument_panels.py and substituted by
`verbs::apply_params` BEFORE any widget value goes in. A placeholder in neither
column is a name nothing will ever fill — the verb refuses to send rather than
half-build the command, so it shows up here as a table bug, not a runtime mystery.

    python3 Deployment/build_yak_command_list.py           # rewrite both files
    python3 Deployment/build_yak_command_list.py --check    # exit 1 if stale
"""
import argparse
import csv
import glob
import json
import os
import re
import sys

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
YAK_ROOT = os.path.join(REPO_ROOT, "BackEnd", "openair-yak", "Yak")
CSV_PATH = os.path.join(YAK_ROOT, "CommandList.csv")
XLSX_PATH = os.path.join(YAK_ROOT, "CommandList.xlsx")

VERBS = ("set", "rig", "nab", "do")
COLUMNS = [
    "Family", "Model", "Verb", "Command", "Description", "SCPI", "SCPI (short)",
    "Arguments", "Arg kind", "Arg values", "Instance params", "Group", "Subsystem",
    "Returns", "Return type", "Return unit", "Return fields",
    "SCPI statements", "Unverified", "File",
]
WIDTHS = [12, 11, 10, 49, 60, 60, 44, 42, 11, 34, 17, 60, 18,
          9, 14, 13, 30, 17, 12, 36]

PLACEHOLDER = re.compile(r"<(\w+)>")


def describe_returns(returns):
    """Flatten a `returns` object into the four reply columns.

    A compound query — `MODE?;MEAS:VOLT?;MEAS:CURR?` — declares one `fields`
    entry per value, and the columns stay parallel so row-wise they read as
    tuples: field n's type is the nth entry of `Return type`. A field that never
    got a type is written `?` rather than blank, because "unknown" and "the
    reply carries no unit" are different facts and a blank cannot say which.
    """
    if not returns:
        return "", "", "", ""
    count = returns.get("count", "")
    fields = returns.get("fields")
    if not fields:
        return count, returns.get("type", ""), returns.get("unit", ""), ""
    return (
        count,
        "; ".join(f.get("type", "?") for f in fields),
        "; ".join(f.get("unit", "") for f in fields),
        "; ".join(f.get("name", "") for f in fields),
    )


def rows():
    """Every command in the tree, as sheet rows, in a stable order.

    Sorted by (family, model, verb, command) so a regeneration after an edit
    produces a diff of the edit rather than of the dict order it happened to be
    written in.
    """
    out = []
    for path in sorted(glob.glob(os.path.join(YAK_ROOT, "*", "*", "commands.json"))):
        rel = os.path.relpath(path, YAK_ROOT).replace(os.sep, "/")
        with open(path) as f:
            table = json.load(f)

        # The declared model wins over the directory. Filing a command under its
        # folder name is exactly the bug the table format was introduced to end.
        family = table.get("family") or rel.split("/")[0]
        model = table.get("model") or rel.split("/")[1]

        for verb in VERBS:
            block = table.get(verb) or {}
            for name, cmd in sorted(block.items()):
                scpi = cmd.get("scpi", "")
                args = cmd.get("args") or []
                # Anything in the template the operator does not supply is a
                # per-instance constant, in template order.
                found = PLACEHOLDER.findall(scpi)
                params = [p for p in dict.fromkeys(found) if p not in args]
                count, rtype, runit, rfields = describe_returns(cmd.get("returns"))
                arg = cmd.get("arg") or {}
                out.append({
                    "Family": family,
                    "Model": model,
                    "Verb": verb.upper(),
                    "Command": name,
                    "Description": cmd.get("description", ""),
                    "SCPI": scpi,
                    "SCPI (short)": cmd.get("scpiFast", ""),
                    "Arguments": "; ".join(args),
                    "Arg kind": arg.get("kind", ""),
                    # An enum's options are the widget: a selector cannot be
                    # authored from the SCPI string alone.
                    "Arg values": "; ".join(arg.get("values") or []),
                    "Instance params": "; ".join(params),
                    "Group": cmd.get("group", ""),
                    "Subsystem": cmd.get("subsystem", ""),
                    "Returns": count,
                    "Return type": rtype,
                    "Return unit": runit,
                    "Return fields": rfields,
                    "SCPI statements": len([s for s in scpi.split(";") if s.strip()]),
                    "Unverified": "yes" if cmd.get("unverified") else "",
                    "File": rel,
                })
    out.sort(key=lambda r: (r["Family"], r["Model"], r["Verb"], r["Command"]))
    return out


def write_csv(data, path):
    with open(path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=COLUMNS)
        w.writeheader()
        w.writerows(data)


def write_xlsx(data, path):
    """Same rows, frozen header and an autofilter. Skipped if openpyxl is absent
    — the CSV is the artifact that matters, the workbook is a convenience."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("   ⚠️  openpyxl not installed — wrote the CSV only")
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "CommandList"
    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in data:
        ws.append([row[c] for c in COLUMNS])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(data) + 1}"
    for i, width in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    wb.save(path)
    return True


def summarize(data):
    print(f"   ✅ {len(data)} commands from "
          f"{len({(r['Family'], r['Model']) for r in data})} models")
    for verb in ("SET", "RIG", "NAB", "DO"):
        n = sum(1 for r in data if r["Verb"] == verb)
        print(f"      {verb:<4} {n:5d}")
    unverified = sum(1 for r in data if r["Unverified"])
    print(f"      unverified {unverified} ({unverified * 100 // max(len(data), 1)}%)")

    # A placeholder that is neither an argument nor an instance param can never
    # be filled: fill_placeholders refuses the command and the control is dead.
    # Cheap to check here, invisible everywhere else.
    stamped = {"chan", "slot"}
    orphans = [r for r in data
               if r["Instance params"]
               and set(r["Instance params"].split("; ")) - stamped]
    if orphans:
        print(f"   ⚠️  {len(orphans)} commands need instance params beyond "
              f"<chan>/<slot> — nothing stamps those, so they cannot send:")
        for r in orphans[:10]:
            print(f"      {r['Family']}/{r['Model']} {r['Command']}: "
                  f"<{r['Instance params']}>")
        if len(orphans) > 10:
            print(f"      … and {len(orphans) - 10} more")


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--check", action="store_true",
                    help="report drift without writing; exit 1 if stale")
    args = ap.parse_args()

    data = rows()

    if args.check:
        try:
            with open(CSV_PATH, newline="") as f:
                current = list(csv.DictReader(f))
        except OSError:
            print("   ❌ no CommandList.csv — run without --check")
            return 1
        fresh = [{k: str(v) for k, v in r.items()} for r in data]
        current = [{k: (v or "") for k, v in r.items()} for r in current]
        if current == fresh:
            print(f"   ✅ CommandList.csv is current ({len(data)} commands)")
            return 0
        print(f"   ❌ CommandList.csv is stale: sheet has {len(current)} rows, "
              f"the tables have {len(data)}")
        return 1

    write_csv(data, CSV_PATH)
    wrote_xlsx = write_xlsx(data, XLSX_PATH)
    print(f"   ✅ wrote {os.path.relpath(CSV_PATH, REPO_ROOT)}"
          + (f" and {os.path.relpath(XLSX_PATH, REPO_ROOT)}" if wrote_xlsx else ""))
    summarize(data)
    return 0


if __name__ == "__main__":
    sys.exit(main())
